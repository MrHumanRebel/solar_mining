#!/usr/bin/env python3
import argparse
import ast
import json
import re
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook
from io import BytesIO


def parse_ymd(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def chunks_inclusive(start: date, end: date, chunk_days: int):
    s = start
    while s <= end:
        e = min(s + timedelta(days=chunk_days - 1), end)
        yield s, e
        s = e + timedelta(days=1)


def sanitize_filename(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", name)


def clipboard_text() -> str:
    try:
        return subprocess.check_output(["pbpaste"], text=True)
    except Exception as e:
        raise RuntimeError(f"Failed to read clipboard via pbpaste: {e}")


def extract_fetch_parts(fetch_text: str):
    m = re.search(r"fetch\(\s*(['\"])(https?://[^'\"]+)\1\s*,\s*\{", fetch_text)
    if not m:
        raise ValueError("Could not find fetch(URL, {...}) in clipboard. Copy as fetch again.")
    url = m.group(2)

    start = fetch_text.find("{", m.end() - 1)
    if start < 0:
        raise ValueError("Could not locate options object '{...}'.")

    depth = 0
    end = None
    for i in range(start, len(fetch_text)):
        ch = fetch_text[i]
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    if end is None:
        raise ValueError("Could not find end of options object. Clipboard seems truncated.")

    obj_text = fetch_text[start : end + 1]
    obj_text_py = (
        obj_text
        .replace(": true", ": True")
        .replace(": false", ": False")
        .replace(": null", ": None")
    )

    options = ast.literal_eval(obj_text_py)
    return url, options


def filter_headers(browser_headers: dict):
    filtered = {}
    for k, v in (browser_headers or {}).items():
        lk = k.lower()
        if lk.startswith("sec-"):
            continue
        if lk in (":authority", ":method", ":path", ":scheme"):
            continue
        filtered[k] = v
    return filtered


def parse_xlsx_to_json(xlsx_bytes: bytes, drop_columns: set[str]):
    """
    Reads first worksheet:
      - row 1 assumed header
      - returns list[dict] rows
    Drops columns whose header matches any in drop_columns.
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []

    header = [("" if h is None else str(h).strip()) for h in header]

    # Determine which indices to keep
    keep_idxs = [i for i, h in enumerate(header) if h and h not in drop_columns]
    keep_headers = [header[i] for i in keep_idxs]

    out = []
    for r in rows:
        if r is None:
            continue
        # skip empty
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in r):
            continue
        obj = {}
        for idx, key in zip(keep_idxs, keep_headers):
            val = r[idx] if idx < len(r) else None
            # JSON serializable conversions
            if isinstance(val, (datetime, date)):
                val = val.isoformat()
            obj[key] = val
        out.append(obj)
    return out


def main():
    ap = argparse.ArgumentParser(description="Copy-as-fetch -> download XLSX -> convert to JSON -> drop Plant Name column")
    ap.add_argument("--from", dest="start", required=True, help="YYYY-MM-DD (e.g. 2025-05-01)")
    ap.add_argument("--to", dest="end", default=date.today().isoformat(), help="YYYY-MM-DD (default: today)")
    ap.add_argument("--dir", default="./solarman_json", help="Output directory for JSON files")
    ap.add_argument("--chunk-days", type=int, default=29, help="Days per request (<=30). Default 29.")
    ap.add_argument("--pretty", action="store_true", help="Pretty-print JSON (bigger files)")
    ap.add_argument("--keep-debug", action="store_true", help="Save .debug for non-XLSX responses")
    args = ap.parse_args()

    start = parse_ymd(args.start)
    end = parse_ymd(args.end)
    if end < start:
        raise SystemExit("ERROR: --to earlier than --from")
    if not (1 <= args.chunk_days <= 30):
        raise SystemExit("ERROR: --chunk-days must be 1..30")

    fetch_text = clipboard_text()
    if "fetch(" not in fetch_text:
        raise SystemExit("ERROR: Clipboard does not look like 'Copy as fetch'. Copy as fetch again first.")

    url, options = extract_fetch_parts(fetch_text)

    headers = options.get("headers") or {}
    method = (options.get("method") or "GET").upper()
    referrer = options.get("referrer")
    body_raw = options.get("body")

    if method != "POST":
        raise SystemExit(f"ERROR: Expected POST. Got method={method}")
    if not body_raw:
        raise SystemExit("ERROR: No body found in fetch snippet.")

    base_payload = json.loads(body_raw) if isinstance(body_raw, str) else body_raw
    timeType = str(base_payload.get("timeType", "1"))
    reportFields = base_payload.get("reportFields")
    if not isinstance(reportFields, list) or not reportFields:
        raise SystemExit("ERROR: reportFields missing/invalid in payload.")

    # Drop Plant Name:
    # - by visible header "Plant Name"
    # - and also storageName NAME (some exports may show "NAME" header)
    drop_headers = {"Plant Name", "NAME"}

    out_dir = Path(args.dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    sess = requests.Session()
    sess.headers.update(filter_headers(headers))
    if referrer:
        sess.headers.setdefault("referer", referrer)
        m = re.match(r"^(https?://[^/]+)", referrer)
        if m:
            sess.headers.setdefault("origin", m.group(1))
    sess.headers.setdefault("user-agent", "Mozilla/5.0")
    sess.headers.setdefault("content-type", "application/json;charset=UTF-8")
    sess.headers.setdefault("accept", "application/json, text/plain, */*")

    ok = 0
    fail = 0

    for s, e in chunks_inclusive(start, end, args.chunk_days):
        payload = {
            "timeType": timeType,
            "startTime": s.isoformat(),
            "endTime": e.isoformat(),
            "reportFields": reportFields,
        }

        key = sanitize_filename(f"{s.isoformat()}_{e.isoformat()}")
        json_path = out_dir / f"{key}.json"
        dbg_path = out_dir / f"{key}.debug"

        print(f"[+] Fetch+convert: {payload['startTime']} -> {payload['endTime']}", flush=True)

        r = sess.post(url, data=json.dumps(payload), timeout=180)
        ct = (r.headers.get("content-type") or "").lower()

        if r.status_code == 200 and "spreadsheetml.sheet" in ct:
            records = parse_xlsx_to_json(r.content, drop_columns=drop_headers)
            with open(json_path, "w", encoding="utf-8") as f:
                if args.pretty:
                    json.dump(records, f, ensure_ascii=False, indent=2)
                else:
                    json.dump(records, f, ensure_ascii=False, separators=(",", ":"))
            ok += 1
            print(f"    [+] Saved JSON: {json_path} ({len(records)} rows)")
            continue

        fail += 1
        print(f"    [!] Failed: HTTP {r.status_code}, ct={ct or 'n/a'}", file=sys.stderr)
        if args.keep_debug:
            try:
                if "application/json" in ct or (r.text or "").strip().startswith("{"):
                    dbg_path.write_text(json.dumps(r.json(), ensure_ascii=False, indent=2), encoding="utf-8")
                else:
                    dbg_path.write_text(r.text or "", encoding="utf-8")
                print(f"    [i] Debug saved: {dbg_path}", file=sys.stderr)
            except Exception:
                pass

    print(f"\n[OK] Done. Wrote {ok} JSON files to: {out_dir.resolve()}")
    if fail:
        print(f"[!] Failed chunks: {fail} (copy a fresh 'Copy as fetch' if token expires).", file=sys.stderr)


if __name__ == "__main__":
    main()